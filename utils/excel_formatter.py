import os
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import logging

logger = logging.getLogger(__name__)


class ExcelFormatter:
    """
    Clase para exportar DataFrames a Excel con formato de tabla.
    """

    def __init__(self):
        self.italic_font = Font(italic=True)
        self.yellow_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        self.light_blue_fill = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")  # Azul claro
        self.center_alignment = Alignment(horizontal='center', vertical='center')

    def _clean_na(self, df):
        """
        Convierte pandas.NA a None para que sea compatible con openpyxl.
        """
        df = df.copy()
        df = df.replace({pd.NA: None})
        df = df.replace({np.nan: None})
        return df

    def _is_subtotal_or_total_row(self, row, idx_descripcion):
        """
        Verifica si una fila es SUBTOTAL o TOTAL basado en la columna 'Descripción'.
        """
        if idx_descripcion is None or idx_descripcion < 0:
            return False
        if idx_descripcion >= len(row):
            return False
        valor = row[idx_descripcion]
        if valor is None:
            return False
        return str(valor).strip().upper() in ['SUBTOTAL', 'TOTAL']

    def export(self, df, excel_path, sheet_name="Hoja1", period="", card_number="9944", horizontal=True):
        """
        Exporta DataFrame a Excel.

        Parameters
        ----------
        df : pandas.DataFrame
            DataFrame con CUALQUIER estructura
        excel_path : str
            Ruta del archivo Excel
        sheet_name : str
            Nombre de la hoja de Excel
        period : str
            Nombre del período para el encabezado
        card_number : str
            Número de tarjeta para el encabezado
        horizontal : bool
            True: Tablas de 40 registros una al lado de la otra
            False: Una sola tabla con todos los registros

        Returns
        -------
        str
            Ruta del archivo generado
        """
        
        # Validar que df tenga datos
        if df is None or df.empty:
            logger.warning(f"DataFrame '{sheet_name}' está vacío, no se exportará")
            return excel_path
        
        # LIMPIAR: Convertir pandas.NA a None para openpyxl
        df = self._clean_na(df)
        
        # Crear o cargar workbook
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
        else:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
        
        # Eliminar hoja si ya existe
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        
        # Crear nueva hoja
        ws = wb.create_sheet(sheet_name)
        
        if horizontal:
            self._write_tables_horizontal(ws, df, period, card_number)
        else:
            self._write_simple_table(ws, df)
        
        wb.save(excel_path)
        logger.info(f"Hoja '{sheet_name}' exportada con {len(df)} registros")
        
        return excel_path

    def _write_tables_horizontal(self, ws, df, period, card_number):
        """
        Escribe tablas de 40 registros una al lado de la otra.
        """
        registros_por_tabla = 40
        columnas = list(df.columns)
        num_columnas = len(columnas)
        espacio_columnas = 1
        
        # Verificar si existe la columna 'Descripción' para SUBTOTAL/TOTAL
        tiene_descripcion = 'Descripción' in df.columns
        idx_descripcion = columnas.index('Descripción') if tiene_descripcion else -1
        
        # Verificar si existe la columna 'Factura' para el fondo amarillo
        tiene_factura = 'Factura' in df.columns
        idx_factura = columnas.index('Factura') if tiene_factura else -1
        
        # Dividir en tablas de 40 registros
        df = df.reset_index(drop=True)
        tablas = []
        for i in range(0, len(df), registros_por_tabla):
            tabla = df.iloc[i:i + registros_por_tabla].copy()
            tablas.append(tabla)
        
        fila_actual = 1
        columna_actual = 1
        
        for idx, tabla in enumerate(tablas, start=1):
            n_registros = len(tabla)
            
            # ============================================================
            # 1. Encabezado de la tabla (fila 1)
            # ============================================================
            encabezado = f"{card_number} - {period}"
            ws.merge_cells(
                start_row=fila_actual, 
                start_column=columna_actual,
                end_row=fila_actual, 
                end_column=columna_actual + num_columnas - 1
            )
            celda = ws.cell(row=fila_actual, column=columna_actual, value=encabezado)
            celda.font = Font(bold=True, size=12)
            celda.alignment = Alignment(horizontal='center', vertical='center')
            
            # ============================================================
            # 2. Subtítulos (fila 2) - TODAS las columnas
            # ============================================================
            for j, col in enumerate(columnas):
                celda = ws.cell(row=fila_actual + 1, column=columna_actual + j, value=col)
                celda.font = Font(bold=True)
                celda.alignment = Alignment(horizontal='center', vertical='center')
            
            # ============================================================
            # 3. Datos (desde fila 3)
            # ============================================================
            for i, row in enumerate(tabla.itertuples(index=False), start=0):
                fila_dato = fila_actual + 2 + i

                # Verificar si esta fila es SUBTOTAL o TOTAL
                es_subtotal_total = self._is_subtotal_or_total_row(row, idx_descripcion)

                for j, valor in enumerate(row):
                    # Asegurar que el valor sea compatible con Excel
                    if pd.isna(valor) or valor is None:
                        valor = None
                    celda = ws.cell(row=fila_dato, column=columna_actual + j, value=valor)
                    celda.font = self.italic_font
                    
                    # ============================================================
                    # Si es SUBTOTAL o TOTAL → FONDO AZUL CLARO en TODA la fila
                    # ============================================================
                    if es_subtotal_total:
                        celda.fill = self.light_blue_fill
                
                # ============================================================
                # Si la columna 'Factura' está vacía → FONDO AMARILLO en TODA la fila
                # (SOLO si NO es SUBTOTAL/TOTAL, para no sobrescribir el azul)
                # ============================================================
                if not es_subtotal_total and tiene_factura and idx_factura >= 0:
                    valor_factura = getattr(row, 'Factura', '')
                    if pd.isna(valor_factura) or valor_factura == '' or str(valor_factura).strip() == '':
                        for k in range(num_columnas):
                            celda_fila = ws.cell(row=fila_dato, column=columna_actual + k)
                            celda_fila.fill = self.yellow_fill
            
            # ============================================================
            # 4. Ajustar ancho de columnas
            # ============================================================
            for j, col in enumerate(columnas):
                col_letter = get_column_letter(columna_actual + j)
                valores = []
                for fila in range(fila_actual, fila_actual + 2 + n_registros):
                    celda = ws.cell(row=fila, column=columna_actual + j)
                    if celda.value is not None:
                        valores.append(str(celda.value))
                ancho_max = max([len(str(col))] + [len(v) for v in valores])
                ws.column_dimensions[col_letter].width = min(ancho_max + 3, 50)
            
            # ============================================================
            # 5. Mover a la siguiente posición (horizontal)
            # ============================================================
            columna_actual += num_columnas + espacio_columnas

    def _write_simple_table(self, ws, df):
        """
        Escribe una sola tabla con TODOS los registros del DataFrame.
        Sin colores, sin divisiones.
        """
        columnas = list(df.columns)
        num_columnas = len(columnas)
        
        # Verificar si existe la columna 'Descripción' para SUBTOTAL/TOTAL
        tiene_descripcion = 'Descripción' in df.columns
        idx_descripcion = columnas.index('Descripción') if tiene_descripcion else -1
        
        # ============================================================
        # 1. Encabezados (fila 1)
        # ============================================================
        for j, col in enumerate(columnas, start=1):
            celda = ws.cell(row=1, column=j, value=col)
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal='center', vertical='center')
        
        # ============================================================
        # 2. Datos (desde fila 2)
        # ============================================================
        for i, row in enumerate(df.itertuples(index=False), start=0):
            fila_dato = i + 2
            
            # Verificar si esta fila es SUBTOTAL o TOTAL
            es_subtotal_total = self._is_subtotal_or_total_row(row, idx_descripcion)
            
            for j, valor in enumerate(row, start=1):
                # Asegurar que el valor sea compatible con Excel
                if pd.isna(valor) or valor is None:
                    valor = None
                celda = ws.cell(row=fila_dato, column=j, value=valor)
                celda.font = self.italic_font
                
                # ============================================================
                # Si es SUBTOTAL o TOTAL → FONDO AZUL CLARO en TODA la fila
                # ============================================================
                if es_subtotal_total:
                    celda.fill = self.light_blue_fill
        
        # ============================================================
        # 3. Ajustar ancho de columnas
        # ============================================================
        for j, col in enumerate(columnas, start=1):
            col_letter = get_column_letter(j)
            valores = []
            for fila in range(1, len(df) + 2):
                celda = ws.cell(row=fila, column=j)
                if celda.value is not None:
                    valores.append(str(celda.value))
            ancho_max = max([len(str(col))] + [len(v) for v in valores])
            ws.column_dimensions[col_letter].width = min(ancho_max + 3, 50)